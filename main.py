import os
import shutil
from copy import copy
from datetime import datetime
import json
import subprocess
import sys
import tempfile
import threading
import tkinter as tk
from tkinter import filedialog
from urllib import error as urllib_error
from urllib import request as urllib_request

import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


DROP_COLUMNS = ["Nome", "Canal", "CAMPANHA", "Questão"]
MAPPING_COLUMNS = {
	"DATA E HORA": "Data",
	"NOME CLIENTE": "Cliente",
	"NÚMERO CLIENTE": "NÚMERO DO CLIENTE",
	"PESQUISA": "Pesquisa",
	"NOTA": "Nota",
	"PROTOCOLO": "PROTOCOLO",
	"DIA": "Data_Correta",
}
REQUIRED_REPORT_COLUMNS = {"Data", "Pesquisa", "Nota"}
ROW_DETECTION_HEADERS = {
	"DATA E HORA",
	"NOME CLIENTE",
	"NÚMERO CLIENTE",
	"PESQUISA",
	"NOTA",
	"PROTOCOLO",
}
APP_EXECUTABLE_NAME = "AtualizadorValidacao.exe"
UPDATE_CONFIG_FILE = "update_config.json"
UPDATE_USER_AGENT = "AtualizadorValidacaoUpdater/1.0"


def get_app_directory() -> str:
	if getattr(sys, "frozen", False):
		return os.path.dirname(os.path.abspath(sys.executable))
	return os.path.dirname(os.path.abspath(__file__))


def load_update_config() -> dict[str, object]:
	config_path = os.path.join(get_app_directory(), UPDATE_CONFIG_FILE)
	if not os.path.isfile(config_path):
		raise RuntimeError(
			"Arquivo update_config.json não encontrado ao lado do executável."
		)

	try:
		with open(config_path, encoding="utf-8") as config_file:
			raw_config = json.load(config_file)
	except json.JSONDecodeError as exc:
		raise RuntimeError("update_config.json está inválido.") from exc

	repo_owner = str(raw_config.get("repo_owner", "")).strip()
	repo_name = str(raw_config.get("repo_name", "")).strip()
	branch = str(raw_config.get("branch", "main")).strip() or "main"
	asset_path = str(raw_config.get("asset_path", "")).strip()
	timeout_seconds_raw = raw_config.get("timeout_seconds", 60)

	if not repo_owner or not repo_name or not asset_path:
		raise RuntimeError(
			"Configure repo_owner, repo_name e asset_path no update_config.json."
		)

	if repo_owner.upper().startswith("SEU_") or repo_name.upper().startswith("SEU_"):
		raise RuntimeError(
			"Atualize repo_owner e repo_name no update_config.json com seu repositório real."
		)

	try:
		timeout_seconds = max(5, int(timeout_seconds_raw))
	except (TypeError, ValueError) as exc:
		raise RuntimeError("timeout_seconds inválido no update_config.json.") from exc

	return {
		"repo_owner": repo_owner,
		"repo_name": repo_name,
		"branch": branch,
		"asset_path": asset_path,
		"timeout_seconds": timeout_seconds,
	}


def resolve_update_download_url(config: dict[str, object]) -> str:
	repo_owner = str(config["repo_owner"])
	repo_name = str(config["repo_name"])
	branch = str(config["branch"])
	asset_path = str(config["asset_path"])
	timeout_seconds = int(config["timeout_seconds"])

	api_url = (
		f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/"
		f"{asset_path}?ref={branch}"
	)
	request = urllib_request.Request(
		api_url,
		headers={
			"Accept": "application/vnd.github+json",
			"User-Agent": UPDATE_USER_AGENT,
		},
	)

	try:
		with urllib_request.urlopen(request, timeout=timeout_seconds) as response:
			payload = json.loads(response.read().decode("utf-8"))
	except urllib_error.HTTPError as exc:
		if exc.code == 404:
			raise RuntimeError(
				"Arquivo de atualização não encontrado no GitHub. Verifique update_config.json."
			) from exc
		if exc.code == 403:
			raise RuntimeError(
				"Acesso negado ao GitHub API. Verifique permissões do repositório."
			) from exc
		raise RuntimeError(f"Falha ao consultar atualização: HTTP {exc.code}.") from exc
	except urllib_error.URLError as exc:
		raise RuntimeError("Falha de rede ao consultar atualização no GitHub.") from exc

	download_url = payload.get("download_url")
	if not isinstance(download_url, str) or not download_url.strip():
		raise RuntimeError("Não foi possível resolver o link de download no GitHub.")

	return download_url


def download_update_executable(
	download_url: str,
	destination_path: str,
	timeout_seconds: int,
) -> None:
	request = urllib_request.Request(
		download_url,
		headers={"User-Agent": UPDATE_USER_AGENT},
	)

	try:
		with urllib_request.urlopen(request, timeout=timeout_seconds) as response:
			with open(destination_path, "wb") as output_file:
				shutil.copyfileobj(response, output_file)
	except urllib_error.URLError as exc:
		raise RuntimeError("Falha de rede durante o download da atualização.") from exc
	except OSError as exc:
		raise RuntimeError("Não foi possível salvar o arquivo de atualização.") from exc

	try:
		with open(destination_path, "rb") as downloaded_file:
			signature = downloaded_file.read(2)
	except OSError as exc:
		raise RuntimeError("Não foi possível validar o arquivo baixado.") from exc

	if signature != b"MZ":
		try:
			os.remove(destination_path)
		except OSError:
			pass
		raise RuntimeError("O arquivo baixado não parece ser um executável válido.")


def create_update_script(target_executable: str, downloaded_executable: str) -> str:
	script_path = os.path.join(tempfile.gettempdir(), "atualizador_validacao_update.bat")
	script_content = "\r\n".join(
		[
			"@echo off",
			"setlocal",
			f"set \"TARGET={target_executable}\"",
			f"set \"SOURCE={downloaded_executable}\"",
			"for /L %%i in (1,1,20) do (",
			"  copy /Y \"%SOURCE%\" \"%TARGET%\" >nul && goto launch",
			"  timeout /t 1 /nobreak >nul",
			")",
			"goto cleanup",
			":launch",
			"start \"\" \"%TARGET%\"",
			":cleanup",
			"del \"%SOURCE%\" >nul 2>nul",
			"del \"%~f0\" >nul 2>nul",
			"endlocal",
		]
	)

	try:
		with open(script_path, "w", encoding="utf-8", newline="\r\n") as script_file:
			script_file.write(script_content)
	except OSError as exc:
		raise RuntimeError("Não foi possível preparar o instalador da atualização.") from exc

	return script_path


def read_report(report_path: str) -> pd.DataFrame:
	extension = os.path.splitext(report_path)[1].lower()
	dtype_map = {"PROTOCOLO": str, "NÚMERO DO CLIENTE": str}

	if extension in {".xls", ".xlsx", ".xlsm"}:
		try:
			return pd.read_excel(report_path, dtype=dtype_map)
		except ImportError as exc:
			if extension == ".xls":
				raise RuntimeError(
					"Falha ao ler .xls. Instale a dependência 'xlrd'."
				) from exc
			raise RuntimeError(
				"Falha ao ler o Excel. Verifique dependências como openpyxl."
			) from exc

	if extension == ".csv":
		try:
			return pd.read_csv(report_path, dtype=dtype_map, sep=None, engine="python")
		except Exception as exc:
			raise RuntimeError(f"Não foi possível ler o CSV: {exc}") from exc

	raise ValueError("Formato de relatório não suportado. Use .xls, .xlsx, .xlsm ou .csv.")


def validate_report_columns(df: pd.DataFrame) -> None:
	missing = sorted(REQUIRED_REPORT_COLUMNS.difference(df.columns))
	if missing:
		raise ValueError(
			"Colunas obrigatórias ausentes no relatório: " + ", ".join(missing)
		)


def transform_report(df: pd.DataFrame, limit: int, max_note: float) -> pd.DataFrame:
	validate_report_columns(df)

	transformed = df.drop(columns=DROP_COLUMNS, errors="ignore").copy()
	transformed["Nota"] = pd.to_numeric(transformed["Nota"], errors="coerce")
	transformed = transformed[transformed["Nota"].notna()]
	transformed = transformed[transformed["Nota"] <= max_note]

	if transformed.empty:
		return transformed

	transformed["Data_Correta"] = pd.to_datetime(
		transformed["Data"], dayfirst=True, errors="coerce"
	)
	transformed = transformed[transformed["Data_Correta"].notna()]

	if transformed.empty:
		return transformed

	transformed["Data_Correta"] = transformed["Data_Correta"].dt.strftime("%d/%m/%Y")

	transformed = (
		transformed.groupby(["Data_Correta", "Pesquisa"], dropna=False)
		.head(limit)
		.reset_index(drop=True)
	)
	return transformed


def get_sheet_names(validation_path: str) -> list[str]:
	workbook = load_workbook(validation_path, read_only=True)
	try:
		return list(workbook.sheetnames)
	finally:
		workbook.close()


def get_sheet_headers(validation_path: str, sheet_name: str) -> list[str]:
	workbook = load_workbook(validation_path)
	try:
		if sheet_name not in workbook.sheetnames:
			raise ValueError(f"A aba '{sheet_name}' não existe no arquivo de validação.")
		worksheet = workbook[sheet_name]
		headers = [cell.value for cell in worksheet[1] if cell.value is not None]
		if not headers:
			raise ValueError(
				f"A primeira linha da aba '{sheet_name}' esta vazia. "
				"Não foi possível identificar cabeçalhos."
			)
		return headers
	finally:
		workbook.close()


def validate_mapping_for_headers(headers: list[str], df: pd.DataFrame) -> None:
	mapped_headers = [header for header in headers if header in MAPPING_COLUMNS]
	if not mapped_headers:
		raise ValueError(
			"A aba selecionada não possui colunas compatíveis com o mapeamento. "
			"Selecione a aba que contém colunas como 'DATA E HORA', 'PESQUISA', 'NOTA' e 'PROTOCOLO'."
		)

	expected_source_columns = {
		MAPPING_COLUMNS[header]
		for header in mapped_headers
	}
	missing = sorted(col for col in expected_source_columns if col not in df.columns)
	if missing:
		raise ValueError(
			"Colunas necessárias para preencher a aba selecionada estão ausentes: "
			+ ", ".join(missing)
		)


def build_rows(df: pd.DataFrame, headers: list[str]) -> list[list[object]]:
	rows: list[list[object]] = []
	for _, row in df.iterrows():
		new_row = []
		for header in headers:
			source_col = MAPPING_COLUMNS.get(header)
			value = row.get(source_col, None) if source_col else None
			new_row.append(value)
		rows.append(new_row)
	return rows


def create_backup(validation_path: str) -> str:
	downloads = os.path.join(os.path.expanduser("~"), "Downloads")
	os.makedirs(downloads, exist_ok=True)

	_, extension = os.path.splitext(validation_path)
	backup_extension = extension if extension else ".xlsx"
	timestamp = datetime.now().strftime("%d-%m-%H-%M")
	backup_file_name = f"validacao_backup_{timestamp}{backup_extension}"
	backup_path = os.path.join(downloads, backup_file_name)

	shutil.copy2(validation_path, backup_path)
	return backup_path


def cell_has_real_value(cell) -> bool:
	value = cell.value
	if value is None:
		return False

	if cell.data_type == "f":
		return False

	if isinstance(value, str):
		text = value.strip()
		if text == "":
			return False
		if text.startswith("="):
			return False

	return True


def find_next_data_row(
	worksheet,
	tracked_columns: list[int],
	start_row: int = 2,
	end_row: int | None = None,
) -> int:
	if end_row is None:
		end_row = worksheet.max_row

	if end_row < start_row:
		return start_row

	columns_to_check = tracked_columns or list(range(1, worksheet.max_column + 1))

	for row_index in range(end_row, start_row - 1, -1):
		for col_index in columns_to_check:
			cell = worksheet.cell(row=row_index, column=col_index)
			if cell_has_real_value(cell):
				return row_index + 1

	return start_row


def find_data_bounds(
	worksheet,
	tracked_columns: list[int],
	start_row: int = 2,
	end_row: int | None = None,
) -> tuple[int | None, int | None]:
	if end_row is None:
		end_row = worksheet.max_row

	if end_row < start_row:
		return None, None

	columns_to_check = tracked_columns or list(range(1, worksheet.max_column + 1))
	first_real_row: int | None = None
	last_real_row: int | None = None

	for row_index in range(start_row, end_row + 1):
		row_has_real_value = False
		for col_index in columns_to_check:
			cell = worksheet.cell(row=row_index, column=col_index)
			if cell_has_real_value(cell):
				row_has_real_value = True
				break

		if row_has_real_value:
			if first_real_row is None:
				first_real_row = row_index
			last_real_row = row_index

	return first_real_row, last_real_row


def row_has_real_data(worksheet, row_index: int, tracked_columns: list[int]) -> bool:
	for col_index in tracked_columns:
		cell = worksheet.cell(row=row_index, column=col_index)
		if cell_has_real_value(cell):
			return True
	return False


def row_has_style_data(worksheet, row_index: int, tracked_columns: list[int]) -> bool:
	columns_to_check = tracked_columns or list(range(1, worksheet.max_column + 1))
	for col_index in columns_to_check:
		if worksheet.cell(row=row_index, column=col_index).has_style:
			return True
	return False


def find_style_template_row(
	worksheet,
	tracked_columns: list[int],
	start_row: int,
	search_end_row: int,
) -> int | None:
	for row_index in range(start_row - 1, 0, -1):
		if row_has_style_data(worksheet, row_index, tracked_columns):
			return row_index

	for row_index in range(start_row + 1, search_end_row + 1):
		if row_has_style_data(worksheet, row_index, tracked_columns):
			return row_index

	return None


def copy_row_format(
	worksheet,
	source_row: int,
	target_row: int,
	max_column: int,
) -> None:
	source_height = worksheet.row_dimensions[source_row].height
	target_height = worksheet.row_dimensions[target_row].height
	if source_height is not None and target_height is None:
		worksheet.row_dimensions[target_row].height = source_height

	for col_index in range(1, max_column + 1):
		source_cell = worksheet.cell(row=source_row, column=col_index)
		target_cell = worksheet.cell(row=target_row, column=col_index)

		if target_cell.has_style:
			continue
		if source_cell.has_style:
			target_cell._style = copy(source_cell._style)


def choose_start_row_for_insertion(
	worksheet,
	tracked_columns: list[int],
	rows_to_insert: int,
	start_row: int = 2,
	end_row: int | None = None,
) -> int:
	first_real_row, last_real_row = find_data_bounds(
		worksheet,
		tracked_columns,
		start_row=start_row,
		end_row=end_row,
	)

	if first_real_row is None or last_real_row is None:
		return start_row

	current_gap_start: int | None = None
	current_gap_size = 0

	for row_index in range(start_row, last_real_row + 1):
		if row_has_real_data(worksheet, row_index, tracked_columns):
			current_gap_start = None
			current_gap_size = 0
			continue

		if current_gap_start is None:
			current_gap_start = row_index
			current_gap_size = 1
		else:
			current_gap_size += 1

		if current_gap_size >= rows_to_insert:
			return current_gap_start

	return last_real_row + 1


def find_best_matching_table(worksheet, headers: list[str]):
	mapped_headers = {header for header in headers if header in MAPPING_COLUMNS}
	if not mapped_headers:
		return None

	best_match = None
	best_score = 0

	for table in worksheet.tables.values():
		min_col, min_row, max_col, max_row = range_boundaries(table.ref)
		table_headers = {
			worksheet.cell(row=min_row, column=column_index).value
			for column_index in range(min_col, max_col + 1)
		}
		score = len(mapped_headers.intersection(table_headers))
		if score > best_score:
			best_score = score
			best_match = (table, min_col, min_row, max_col, max_row)

	return best_match if best_score > 0 else None


def append_rows_to_sheet(
	validation_path: str,
	sheet_name: str,
	rows: list[list[object]],
	headers: list[str],
) -> dict[str, int]:
	workbook = load_workbook(validation_path)
	try:
		if sheet_name not in workbook.sheetnames:
			raise ValueError(f"A aba '{sheet_name}' não existe no arquivo de validação.")

		worksheet = workbook[sheet_name]
		tracked_columns = [
			index + 1
			for index, header in enumerate(headers)
			if header in ROW_DETECTION_HEADERS
		]
		if not tracked_columns:
			tracked_columns = [
				index + 1
				for index, header in enumerate(headers)
				if header in MAPPING_COLUMNS
			]

		table_match = find_best_matching_table(worksheet, headers)
		if table_match is None:
			start_row = choose_start_row_for_insertion(
				worksheet,
				tracked_columns,
				rows_to_insert=len(rows),
			)
			style_columns = tracked_columns
			style_scan_end = worksheet.max_row
		else:
			table, min_col, min_row, max_col, max_row = table_match
			table_tracked_columns = [
				col for col in tracked_columns if min_col <= col <= max_col
			]
			if not table_tracked_columns:
				table_tracked_columns = list(range(min_col, max_col + 1))

			start_row = choose_start_row_for_insertion(
				worksheet,
				table_tracked_columns,
				rows_to_insert=len(rows),
				start_row=min_row + 1,
				end_row=max_row,
			)
			style_columns = table_tracked_columns
			style_scan_end = max_row

		style_template_row = find_style_template_row(
			worksheet,
			style_columns,
			start_row=start_row,
			search_end_row=max(style_scan_end, worksheet.max_row),
		)
		max_style_column = max(worksheet.max_column, len(headers))

		for row_index, row in enumerate(rows):
			target_row = start_row + row_index
			if style_template_row is not None:
				copy_row_format(
					worksheet,
					source_row=style_template_row,
					target_row=target_row,
					max_column=max_style_column,
				)

			for col_index, value in enumerate(row, start=1):
				worksheet.cell(row=target_row, column=col_index, value=value)

		if table_match is not None:
			table, min_col, min_row, max_col, max_row = table_match
			inserted_last_row = start_row + len(rows) - 1
			if inserted_last_row > max_row:
				table.ref = (
					f"{get_column_letter(min_col)}{min_row}:"
					f"{get_column_letter(max_col)}{inserted_last_row}"
				)

		workbook.save(validation_path)
	finally:
		workbook.close()

	return {
		"inserted": len(rows),
		"start_row": start_row,
		"end_row": start_row + len(rows) - 1,
	}


def run_pipeline(
	report_path: str,
	validation_path: str,
	sheet_name: str,
	limit: int,
	max_note: float,
) -> dict[str, object]:
	report_df = read_report(report_path)
	total_read = len(report_df)

	transformed_df = transform_report(report_df, limit=limit, max_note=max_note)
	total_filtered = len(transformed_df)

	headers = get_sheet_headers(validation_path, sheet_name)

	if transformed_df.empty:
		return {
			"total_read": total_read,
			"total_filtered": total_filtered,
			"inserted": 0,
			"start_row": None,
			"end_row": None,
			"backup_path": None,
			"sheet_name": sheet_name,
		}

	validate_mapping_for_headers(headers, transformed_df)

	formatted_rows = build_rows(transformed_df, headers)
	backup_path = create_backup(validation_path)
	write_result = append_rows_to_sheet(validation_path, sheet_name, formatted_rows, headers)

	return {
		"total_read": total_read,
		"total_filtered": total_filtered,
		"inserted": write_result["inserted"],
		"start_row": write_result["start_row"],
		"end_row": write_result["end_row"],
		"backup_path": backup_path,
		"sheet_name": sheet_name,
	}


class ArrowStepper(ctk.CTkFrame):
	def __init__(
		self,
		master,
		*,
		initial_value: float,
		step: float,
		min_value: float,
		max_value: float,
		decimals: int,
		integer_mode: bool = False,
		on_change=None,
	) -> None:
		super().__init__(master, fg_color="#0e131b", corner_radius=12, border_width=1, border_color="#2b3342")
		self.step = step
		self.min_value = min_value
		self.max_value = max_value
		self.decimals = decimals
		self.integer_mode = integer_mode
		self.on_change = on_change

		self.value_var = tk.StringVar()

		self.grid_columnconfigure(0, weight=1)

		self.entry = ctk.CTkEntry(
			self,
			textvariable=self.value_var,
			height=36,
			border_width=0,
			fg_color="transparent",
			justify="center",
			font=ctk.CTkFont(size=14, weight="bold"),
			text_color="#e5e7eb",
		)
		self.entry.grid(row=0, column=0, sticky="ew", padx=(10, 2), pady=6)
		self.entry.bind("<FocusOut>", self._normalize_from_entry)
		self.entry.bind("<Return>", self._normalize_from_entry)

		button_column = ctk.CTkFrame(self, fg_color="transparent")
		button_column.grid(row=0, column=1, sticky="ns", padx=(2, 8), pady=6)

		self.up_button = ctk.CTkButton(
			button_column,
			text="▲",
			width=28,
			height=14,
			corner_radius=8,
			fg_color="#2d3646",
			hover_color="#3e495f",
			command=lambda: self._step(+1),
		)
		self.up_button.grid(row=0, column=0, pady=(0, 2))

		self.down_button = ctk.CTkButton(
			button_column,
			text="▼",
			width=28,
			height=14,
			corner_radius=8,
			fg_color="#2d3646",
			hover_color="#3e495f",
			command=lambda: self._step(-1),
		)
		self.down_button.grid(row=1, column=0)

		self.set_value(initial_value, trigger=False)

	def _format_value(self, value: float) -> str:
		if self.integer_mode:
			return str(int(round(value)))
		return f"{value:.{self.decimals}f}"

	def _parse_value(self, text: str) -> float:
		clean = text.strip().replace(",", ".")
		if clean == "":
			raise ValueError("Valor vazio")
		value = float(clean)
		if self.integer_mode:
			value = float(int(round(value)))
		return value

	def _clamp(self, value: float) -> float:
		value = max(self.min_value, min(self.max_value, value))
		if self.integer_mode:
			value = float(int(round(value)))
		return value

	def _step(self, direction: int) -> None:
		try:
			current = self._parse_value(self.value_var.get())
		except ValueError:
			current = self.min_value
		new_value = self._clamp(current + (self.step * direction))
		self.set_value(new_value)

	def _normalize_from_entry(self, _event=None) -> None:
		try:
			value = self._parse_value(self.value_var.get())
		except ValueError:
			value = self.min_value
		self.set_value(self._clamp(value))

	def set_value(self, value: float, trigger: bool = True) -> None:
		value = self._clamp(value)
		self.value_var.set(self._format_value(value))
		if trigger and self.on_change:
			self.on_change()

	def get_value(self) -> float:
		value = self._parse_value(self.value_var.get())
		return self._clamp(value)


class ValidationApp:
	def __init__(self, root: ctk.CTk) -> None:
		self.root = root
		self.root.title("Atualizador Validação")
		self.root.geometry("980x620")
		self.root.minsize(940, 600)
		self.root.configure(fg_color="#0b0d11")

		self.report_path_var = tk.StringVar()
		self.validation_path_var = tk.StringVar()
		self.sheet_var = tk.StringVar(value="")
		self._update_in_progress = False

		self.sheet_menu: ctk.CTkOptionMenu
		self.execute_button: ctk.CTkButton
		self.update_button: ctk.CTkButton
		self.progress: ctk.CTkProgressBar
		self.status_chip: ctk.CTkLabel
		self.status_text: ctk.CTkLabel
		self.summary_box: ctk.CTkTextbox

		self._build_ui()
		self._set_summary("Selecione o relatório, o arquivo de validação e a aba de destino.")
		self._set_status("Aguardando arquivos", tone="neutral")
		self._update_execute_state()

	def _build_ui(self) -> None:
		wrapper = ctk.CTkFrame(self.root, fg_color="transparent")
		wrapper.pack(fill="both", expand=True, padx=18, pady=18)
		wrapper.grid_columnconfigure(0, weight=3)
		wrapper.grid_columnconfigure(1, weight=2)
		wrapper.grid_rowconfigure(1, weight=1)

		header = ctk.CTkFrame(
			wrapper,
			fg_color="#121722",
			corner_radius=18,
			border_width=1,
			border_color="#2a3140",
		)
		header.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 14))
		header.grid_columnconfigure(0, weight=1)

		ctk.CTkLabel(
			header,
			text="Atualizador Validação",
			font=ctk.CTkFont(size=24, weight="bold"),
			text_color="#f3f4f6",
		).grid(row=0, column=0, sticky="w", padx=18, pady=(14, 2))

		ctk.CTkLabel(
			header,
			text="Aplicativo interno para atualizar registros com segurança na planilha de validação.",
			font=ctk.CTkFont(size=13),
			text_color="#9da7b6",
		).grid(row=1, column=0, sticky="w", padx=18, pady=(0, 14))

		self.status_chip = ctk.CTkLabel(
			header,
			text="Aguardando",
			fg_color="#1f2937",
			corner_radius=999,
			text_color="#cbd5e1",
			font=ctk.CTkFont(size=12, weight="bold"),
		)
		self.status_chip.grid(row=0, column=1, rowspan=2, sticky="e", padx=18)

		controls = ctk.CTkFrame(
			wrapper,
			fg_color="#121822",
			corner_radius=18,
			border_width=1,
			border_color="#2a3140",
		)
		controls.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
		controls.grid_columnconfigure(0, weight=1)

		self._build_file_picker(
			controls,
			row=0,
			label="Relatório",
			path_var=self.report_path_var,
			button_label="Selecionar arquivo",
			button_command=self._choose_report,
		)

		self._build_file_picker(
			controls,
			row=2,
			label="Arquivo de validação",
			path_var=self.validation_path_var,
			button_label="Selecionar validação",
			button_command=self._choose_validation,
		)

		ctk.CTkLabel(
			controls,
			text="Aba de destino",
			font=ctk.CTkFont(size=12, weight="bold"),
			text_color="#d5d9e0",
		).grid(row=4, column=0, sticky="w", padx=16, pady=(2, 6))

		self.sheet_menu = ctk.CTkOptionMenu(
			controls,
			values=[""],
			variable=self.sheet_var,
			height=38,
			fg_color="#2d3646",
			button_color="#3a455a",
			button_hover_color="#4a5670",
			dropdown_fg_color="#151b27",
			dropdown_hover_color="#253044",
			text_color="#f3f4f6",
			command=lambda _value: self._update_execute_state(),
		)
		self.sheet_menu.grid(row=5, column=0, sticky="ew", padx=16, pady=(0, 14))

		number_grid = ctk.CTkFrame(controls, fg_color="transparent")
		number_grid.grid(row=6, column=0, sticky="ew", padx=16)
		number_grid.grid_columnconfigure(0, weight=1)
		number_grid.grid_columnconfigure(1, weight=1)

		limit_card = ctk.CTkFrame(number_grid, fg_color="#0f141d", corner_radius=14, border_width=1, border_color="#293142")
		limit_card.grid(row=0, column=0, sticky="ew", padx=(0, 6))
		limit_card.grid_columnconfigure(0, weight=1)

		ctk.CTkLabel(
			limit_card,
			text="Limite por Campanha/Dia",
			font=ctk.CTkFont(size=12, weight="bold"),
			text_color="#cbd5e1",
		).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 6))

		self.limit_stepper = ArrowStepper(
			limit_card,
			initial_value=10,
			step=1,
			min_value=1,
			max_value=9999,
			decimals=0,
			integer_mode=True,
			on_change=self._update_execute_state,
		)
		self.limit_stepper.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))

		note_card = ctk.CTkFrame(number_grid, fg_color="#0f141d", corner_radius=14, border_width=1, border_color="#293142")
		note_card.grid(row=0, column=1, sticky="ew", padx=(6, 0))
		note_card.grid_columnconfigure(0, weight=1)

		ctk.CTkLabel(
			note_card,
			text="Nota máxima",
			font=ctk.CTkFont(size=12, weight="bold"),
			text_color="#cbd5e1",
		).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 6))

		self.max_note_stepper = ArrowStepper(
			note_card,
			initial_value=3,
			step=0.5,
			min_value=0,
			max_value=10,
			decimals=1,
			integer_mode=False,
			on_change=self._update_execute_state,
		)
		self.max_note_stepper.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))

		self.execute_button = ctk.CTkButton(
			controls,
			text="Executar atualização",
			height=46,
			corner_radius=14,
			fg_color="#3d475a",
			hover_color="#4f5c74",
			text_color="#f8fafc",
			font=ctk.CTkFont(size=14, weight="bold"),
			command=self._execute,
		)
		self.execute_button.grid(row=7, column=0, sticky="ew", padx=16, pady=(14, 8))

		self.update_button = ctk.CTkButton(
			controls,
			text="Atualizar aplicativo",
			height=42,
			corner_radius=12,
			fg_color="#2f6d44",
			hover_color="#3f8456",
			text_color="#f8fafc",
			font=ctk.CTkFont(size=13, weight="bold"),
			command=self._update_application,
		)
		self.update_button.grid(row=8, column=0, sticky="ew", padx=16, pady=(0, 8))

		self.progress = ctk.CTkProgressBar(
			controls,
			mode="indeterminate",
			height=8,
			fg_color="#1f2735",
			progress_color="#9ca3af",
		)
		self.progress.grid(row=9, column=0, sticky="ew", padx=16, pady=(0, 8))
		self.progress.grid_remove()

		self.status_text = ctk.CTkLabel(
			controls,
			text="",
			text_color="#a9b4c4",
			font=ctk.CTkFont(size=12),
			justify="left",
			wraplength=560,
		)
		self.status_text.grid(row=10, column=0, sticky="w", padx=16, pady=(0, 14))

		side_panel = ctk.CTkFrame(
			wrapper,
			fg_color="#111720",
			corner_radius=18,
			border_width=1,
			border_color="#2a3140",
		)
		side_panel.grid(row=1, column=1, sticky="nsew")
		side_panel.grid_columnconfigure(0, weight=1)

		ctk.CTkLabel(
			side_panel,
			text="Resumo da execução",
			font=ctk.CTkFont(size=16, weight="bold"),
			text_color="#f3f4f6",
		).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 6))

		self.summary_box = ctk.CTkTextbox(
			side_panel,
			height=260,
			corner_radius=12,
			fg_color="#0c1118",
			border_width=1,
			border_color="#273041",
			text_color="#d9e1ea",
			font=ctk.CTkFont(size=12),
		)
		self.summary_box.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 10))
		self.summary_box.configure(state="disabled")

		side_panel.grid_rowconfigure(1, weight=1)

	def _build_file_picker(
		self,
		parent,
		*,
		row: int,
		label: str,
		path_var: tk.StringVar,
		button_label: str,
		button_command,
	) -> None:
		ctk.CTkLabel(
			parent,
			text=label,
			font=ctk.CTkFont(size=12, weight="bold"),
			text_color="#d5d9e0",
		).grid(row=row, column=0, sticky="w", padx=16, pady=(14, 6))

		row_frame = ctk.CTkFrame(parent, fg_color="transparent")
		row_frame.grid(row=row + 1, column=0, sticky="ew", padx=16)
		row_frame.grid_columnconfigure(0, weight=1)

		entry = ctk.CTkEntry(
			row_frame,
			textvariable=path_var,
			height=38,
			fg_color="#0d121a",
			border_color="#2f394b",
			text_color="#e5e7eb",
		)
		entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
		entry.bind("<Key>", lambda _event: "break")

		ctk.CTkButton(
			row_frame,
			text=button_label,
			width=150,
			height=38,
			corner_radius=12,
			fg_color="#2f394b",
			hover_color="#3f4b61",
			text_color="#f1f5f9",
			command=button_command,
		).grid(row=0, column=1)

	def _set_summary(self, text: str) -> None:
		self.summary_box.configure(state="normal")
		self.summary_box.delete("1.0", "end")
		self.summary_box.insert("1.0", text)
		self.summary_box.configure(state="disabled")

	def _refresh_summary_preview(self, extra: str = "") -> None:
		report_path = self.report_path_var.get().strip() or "(não selecionado)"
		validation_path = self.validation_path_var.get().strip() or "(não selecionado)"
		sheet_name = self.sheet_var.get().strip() or "(não selecionada)"

		preview_lines = [
			"Configuração atual",
			"",
			f"Relatório: {report_path}",
			f"Validação: {validation_path}",
			f"Aba: {sheet_name}",
			f"Limite: {self._safe_limit_preview()}",
			f"Nota máxima: {self._safe_note_preview()}",
		]

		if extra:
			preview_lines.extend(["", extra])

		self._set_summary("\n".join(preview_lines))

	def _safe_limit_preview(self) -> str:
		try:
			return str(self._parse_limit())
		except ValueError:
			return "valor inválido"

	def _safe_note_preview(self) -> str:
		try:
			return f"{self._parse_max_note():.1f}"
		except ValueError:
			return "valor inválido"

	def _set_status(self, text: str, *, tone: str) -> None:
		palette = {
			"neutral": ("#1f2937", "#cbd5e1"),
			"ready": ("#1f3b2d", "#86efac"),
			"warn": ("#3f2f1b", "#fbbf24"),
			"busy": ("#1d334a", "#93c5fd"),
			"error": ("#4b1d1d", "#fca5a5"),
		}
		fg_color, text_color = palette.get(tone, palette["neutral"])
		self.status_chip.configure(text=text, fg_color=fg_color, text_color=text_color)

	def _set_processing(self, value: bool) -> None:
		if value:
			self.progress.grid()
			self.progress.start()
		else:
			self.progress.stop()
			self.progress.set(0)
			self.progress.grid_remove()

	def _update_application(self) -> None:
		if not getattr(sys, "frozen", False):
			self._set_status("Somente no .exe", tone="warn")
			self.status_text.configure(
				text="O botão de atualização fica disponível somente no executável compilado."
			)
			return

		self._update_in_progress = True
		self.execute_button.configure(state="disabled")
		self.update_button.configure(state="disabled")
		self._set_processing(True)
		self._set_status("Atualizando", tone="busy")
		self.status_text.configure(text="Baixando a última versão publicada no GitHub...")
		self._refresh_summary_preview(extra="ATUALIZADOR\nBuscando atualização no GitHub...")

		threading.Thread(target=self._update_worker, daemon=True).start()

	def _update_worker(self) -> None:
		try:
			config = load_update_config()
			download_url = resolve_update_download_url(config)
			timeout_seconds = int(config["timeout_seconds"])
			temp_executable = os.path.join(
				tempfile.gettempdir(),
				f"{os.path.splitext(APP_EXECUTABLE_NAME)[0]}_update.exe",
			)
			download_update_executable(download_url, temp_executable, timeout_seconds)
			target_executable = os.path.abspath(sys.executable)
			update_script = create_update_script(target_executable, temp_executable)
		except Exception as exc:
			self.root.after(0, lambda: self._handle_update_failure(str(exc)))
			return

		self.root.after(0, lambda: self._finalize_update(update_script))

	def _handle_update_failure(self, error_message: str) -> None:
		self._update_in_progress = False
		self._set_processing(False)
		self._set_status("Falha na atualização", tone="error")
		self.status_text.configure(text=error_message)
		self._refresh_summary_preview(extra=f"ATUALIZADOR\n{error_message}")
		self._update_execute_state()

	def _finalize_update(self, update_script: str) -> None:
		try:
			subprocess.Popen(
				["cmd", "/c", update_script],
				creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
			)
		except Exception as exc:
			self._handle_update_failure(f"Não foi possível iniciar o atualizador: {exc}")
			return

		self._set_processing(False)
		self._set_status("Reiniciando", tone="busy")
		self.status_text.configure(
			text="Nova versão baixada com sucesso. O aplicativo será reiniciado."
		)
		self._refresh_summary_preview(
			extra="ATUALIZADOR\nNova versão baixada. Reiniciando aplicativo..."
		)
		self.root.after(900, self.root.destroy)

	def _choose_report(self) -> None:
		path = filedialog.askopenfilename(
			title="Selecione o relatório",
			filetypes=[
				("Arquivos de dados", "*.xls *.xlsx *.xlsm *.csv"),
				("Todos os arquivos", "*.*"),
			],
		)
		if path:
			self.report_path_var.set(path)
			self._set_status("Relatório carregado", tone="neutral")
			self.status_text.configure(text="Arquivo de relatório pronto para validação.")

		self._refresh_summary_preview()
		self._update_execute_state()

	def _choose_validation(self) -> None:
		path = filedialog.askopenfilename(
			title="Selecione o arquivo de validação",
			filetypes=[
				("Excel", "*.xlsx *.xlsm"),
				("Todos os arquivos", "*.*"),
			],
		)
		if not path:
			self._update_execute_state()
			return

		self.validation_path_var.set(path)

		try:
			sheet_names = get_sheet_names(path)
		except Exception as exc:
			self.validation_path_var.set("")
			self.sheet_var.set("")
			self.sheet_menu.configure(values=[""])
			self._set_status("Erro de leitura", tone="error")
			self.status_text.configure(text="Selecione um arquivo de validação válido.")
			self._refresh_summary_preview(extra=f"ERRO: Não foi possível carregar as abas.\n{exc}")
			self._update_execute_state()
			return

		if not sheet_names:
			self.validation_path_var.set("")
			self.sheet_var.set("")
			self.sheet_menu.configure(values=[""])
			self._set_status("Sem abas", tone="error")
			self.status_text.configure(text="Arquivo sem estrutura de abas utilizável.")
			self._refresh_summary_preview(extra="ERRO: O arquivo de validação não possui abas.")
			self._update_execute_state()
			return

		self.sheet_menu.configure(values=sheet_names)
		default_sheet = "Base" if "Base" in sheet_names else sheet_names[0]
		self.sheet_var.set(default_sheet)
		self._set_status("Validação carregada", tone="neutral")
		self.status_text.configure(text=f"Arquivo carregado com {len(sheet_names)} aba(s).")
		self._refresh_summary_preview()
		self._update_execute_state()

	def _parse_limit(self) -> int:
		value = self.limit_stepper.get_value()
		value_int = int(round(value))
		if value_int <= 0:
			raise ValueError("O limite deve ser maior que zero.")
		return value_int

	def _parse_max_note(self) -> float:
		value = self.max_note_stepper.get_value()
		if value < 0:
			raise ValueError("A nota máxima não pode ser negativa.")
		return float(value)

	def _validate_inputs(self) -> tuple[bool, str]:
		report_path = self.report_path_var.get().strip()
		validation_path = self.validation_path_var.get().strip()
		sheet_name = self.sheet_var.get().strip()

		if not report_path:
			return False, "Selecione o relatório."
		if not os.path.isfile(report_path):
			return False, "O relatório selecionado não existe mais."

		if not validation_path:
			return False, "Selecione o arquivo de validação."
		if not os.path.isfile(validation_path):
			return False, "O arquivo de validação selecionado não existe mais."

		if not sheet_name:
			return False, "Selecione a aba de destino."

		try:
			self._parse_limit()
			self._parse_max_note()
		except ValueError as exc:
			return False, str(exc)

		return True, "Pronto para executar."

	def _update_execute_state(self) -> None:
		if self._update_in_progress:
			return

		is_valid, message = self._validate_inputs()
		self.execute_button.configure(state="normal" if is_valid else "disabled")
		self.update_button.configure(
			state="normal" if getattr(sys, "frozen", False) else "disabled"
		)
		if is_valid:
			self._set_status("Pronto", tone="ready")
		else:
			self._set_status("Aguardando ajustes", tone="warn")
		self.status_text.configure(text=message)

	def _execute(self) -> None:
		is_valid, message = self._validate_inputs()
		if not is_valid:
			self._set_status("Ajuste necessário", tone="warn")
			self.status_text.configure(text=message)
			self._refresh_summary_preview(extra=f"VALIDAÇÃO: {message}")
			self._update_execute_state()
			return

		report_path = self.report_path_var.get().strip()
		validation_path = self.validation_path_var.get().strip()
		sheet_name = self.sheet_var.get().strip()
		limit = self._parse_limit()
		max_note = self._parse_max_note()

		self.execute_button.configure(state="disabled")
		self.update_button.configure(state="disabled")
		self._set_processing(True)
		self._set_status("Processando", tone="busy")
		self.status_text.configure(text="Executando a atualização e preparando a inserção na aba selecionada...")
		self.root.update_idletasks()

		try:
			summary = run_pipeline(
				report_path=report_path,
				validation_path=validation_path,
				sheet_name=sheet_name,
				limit=limit,
				max_note=max_note,
			)
		except Exception as exc:
			self._set_processing(False)
			self._set_status("Falha", tone="error")
			self.status_text.configure(text="Falha durante a execução. Verifique a mensagem de erro.")
			self._refresh_summary_preview(extra=f"ERRO: {exc}")
			self._update_execute_state()
			return

		self._set_processing(False)

		inserted = summary["inserted"]
		total_read = summary["total_read"]
		total_filtered = summary["total_filtered"]
		start_row = summary["start_row"]
		end_row = summary["end_row"]
		backup_path = summary["backup_path"]

		if inserted == 0:
			self._set_status("Sem inserções", tone="warn")
			self.status_text.configure(text="Nenhum registro atendeu ao filtro atual.")
			self._refresh_summary_preview(
				extra=(
					"ÚLTIMA EXECUÇÃO\n"
					f"Total lido: {total_read}\n"
					f"Total após filtros: {total_filtered}\n"
					"Nenhuma inserção realizada."
				)
			)
			self._update_execute_state()
			return

		self._set_status("Concluído", tone="ready")
		self.status_text.configure(text="Atualização finalizada com sucesso, mantendo a estrutura da planilha.")
		self._refresh_summary_preview(
			extra=(
				"ÚLTIMA EXECUÇÃO\n"
				f"Total lido: {total_read}\n"
				f"Total após filtros: {total_filtered}\n"
				f"Total inserido: {inserted}\n"
				f"Linhas inseridas: {start_row} até {end_row}\n"
				f"Backup: {backup_path}"
			)
		)

		self._update_execute_state()


def main() -> None:
	ctk.set_appearance_mode("dark")
	ctk.set_default_color_theme("dark-blue")
	root = ctk.CTk()
	app = ValidationApp(root)
	app.root.mainloop()


if __name__ == "__main__":
	main()